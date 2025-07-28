"""
Outils de génération de rapports pour l'ADK (Agent Development Kit)
Permet aux agents de créer des rapports PDF, CSV et autres formats
"""

import os
import json
import csv
from datetime import datetime, timedelta
from typing import Dict, Any, List, Optional
import tempfile
from pathlib import Path
from adk.api.tool_manager import tool

# Import conditionnel des bibliothèques de génération PDF
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    from fpdf import FPDF
    FPDF_AVAILABLE = True
except ImportError:
    FPDF_AVAILABLE = False

# Styles pour les rapports
REPORT_STYLES = {
    "title": {
        "fontSize": 16,
        "fontName": "Helvetica-Bold",
        "alignment": TA_CENTER,
        "spaceAfter": 20
    },
    "heading": {
        "fontSize": 14,
        "fontName": "Helvetica-Bold",
        "alignment": TA_LEFT,
        "spaceAfter": 12
    },
    "normal": {
        "fontSize": 10,
        "fontName": "Helvetica",
        "alignment": TA_LEFT,
        "spaceAfter": 6
    }
}

@tool
async def generate_order_summary_pdf(
    orders: List[Dict[str, Any]],
    report_title: str = "Rapport des Ordres de Travail",
    include_charts: bool = True,
    output_format: str = "pdf"
) -> str:
    """
    Génère un rapport PDF résumé des ordres de travail.
    Utilisez cet outil pour créer des rapports de maintenance professionnels.
    
    Args:
        orders: Liste des ordres de travail avec leurs détails
        report_title: Titre du rapport
        include_charts: Inclure des graphiques (si supporté)
        output_format: Format de sortie ('pdf', 'csv', 'json')
    
    Returns:
        Chemin vers le fichier de rapport généré
    """
    try:
        if not orders:
            return "❌ Aucun ordre de travail fourni pour le rapport"
        
        # Création du nom de fichier
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"rapport_ordres_{timestamp}"
        
        if output_format == "pdf":
            if not REPORTLAB_AVAILABLE:
                return "❌ ReportLab non installé. Installez-le avec: pip install reportlab"
            
            return await _generate_pdf_report(orders, report_title, filename, include_charts)
        
        elif output_format == "csv":
            return await _generate_csv_report(orders, filename)
        
        elif output_format == "json":
            return await _generate_json_report(orders, filename)
        
        else:
            return f"❌ Format non supporté: {output_format}"
            
    except Exception as e:
        return f"❌ Erreur lors de la génération du rapport: {str(e)}"

async def _generate_pdf_report(orders: List[Dict[str, Any]], title: str, filename: str, include_charts: bool) -> str:
    """Génère un rapport PDF avec ReportLab"""
    try:
        filepath = f"reports/{filename}.pdf"
        os.makedirs("reports", exist_ok=True)
        
        doc = SimpleDocTemplate(filepath, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Titre du rapport
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER
        )
        story.append(Paragraph(title, title_style))
        story.append(Spacer(1, 20))
        
        # Informations générales
        story.append(Paragraph("Informations Générales", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        # Statistiques
        total_orders = len(orders)
        urgent_orders = len([o for o in orders if o.get('priority', '').lower() in ['urgent', 'high', 'très élevé']])
        completed_orders = len([o for o in orders if o.get('status', '').lower() in ['completed', 'terminé', 'fini']])
        
        stats_data = [
            ['Métrique', 'Valeur'],
            ['Total des ordres', str(total_orders)],
            ['Ordres urgents', str(urgent_orders)],
            ['Ordres terminés', str(completed_orders)],
            ['Taux de completion', f"{(completed_orders/total_orders*100):.1f}%" if total_orders > 0 else "0%"]
        ]
        
        stats_table = Table(stats_data, colWidths=[2*inch, 1.5*inch])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(stats_table)
        story.append(Spacer(1, 20))
        
        # Détail des ordres
        story.append(Paragraph("Détail des Ordres de Travail", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        # En-têtes du tableau
        headers = ['ID', 'Description', 'Équipement', 'Priorité', 'Statut', 'Date Création']
        table_data = [headers]
        
        for order in orders:
            row = [
                str(order.get('id', 'N/A')),
                str(order.get('description', 'N/A'))[:30] + '...' if len(str(order.get('description', ''))) > 30 else str(order.get('description', 'N/A')),
                str(order.get('equipment_id', 'N/A')),
                str(order.get('priority', 'N/A')),
                str(order.get('status', 'N/A')),
                str(order.get('created_date', 'N/A'))
            ]
            table_data.append(row)
        
        # Tableau des ordres
        order_table = Table(table_data, colWidths=[0.8*inch, 2.5*inch, 1*inch, 0.8*inch, 1*inch, 1.2*inch])
        order_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        story.append(order_table)
        story.append(Spacer(1, 20))
        
        # Pied de page
        footer_text = f"Rapport généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        story.append(Paragraph(footer_text, styles['Normal']))
        
        # Génération du PDF
        doc.build(story)
        
        return f"✅ Rapport PDF généré avec succès: {filepath}"
        
    except Exception as e:
        return f"❌ Erreur lors de la génération du PDF: {str(e)}"

async def _generate_csv_report(orders: List[Dict[str, Any]], filename: str) -> str:
    """Génère un rapport CSV"""
    try:
        filepath = f"reports/{filename}.csv"
        os.makedirs("reports", exist_ok=True)
        
        if not orders:
            return "❌ Aucun ordre à exporter"
        
        # Définition des colonnes
        fieldnames = ['id', 'description', 'equipment_id', 'priority', 'status', 'created_date', 'planned_start', 'planned_end']
        
        with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            
            for order in orders:
                # Nettoyage et formatage des données
                clean_order = {}
                for field in fieldnames:
                    clean_order[field] = str(order.get(field, ''))
                writer.writerow(clean_order)
        
        return f"✅ Rapport CSV généré avec succès: {filepath}"
        
    except Exception as e:
        return f"❌ Erreur lors de la génération du CSV: {str(e)}"

async def _generate_json_report(orders: List[Dict[str, Any]], filename: str) -> str:
    """Génère un rapport JSON"""
    try:
        filepath = f"reports/{filename}.json"
        os.makedirs("reports", exist_ok=True)
        
        report_data = {
            "metadata": {
                "generated_at": datetime.now().isoformat(),
                "total_orders": len(orders),
                "report_type": "maintenance_orders"
            },
            "orders": orders
        }
        
        with open(filepath, 'w', encoding='utf-8') as jsonfile:
            json.dump(report_data, jsonfile, indent=2, ensure_ascii=False)
        
        return f"✅ Rapport JSON généré avec succès: {filepath}"
        
    except Exception as e:
        return f"❌ Erreur lors de la génération du JSON: {str(e)}"

@tool
async def generate_equipment_maintenance_report(
    equipment_data: List[Dict[str, Any]],
    report_period: str = "30",
    include_history: bool = True
) -> str:
    """
    Génère un rapport de maintenance par équipement.
    Utilisez cet outil pour analyser les performances de maintenance des équipements.
    
    Args:
        equipment_data: Données des équipements avec leur historique de maintenance
        report_period: Période d'analyse en jours
        include_history: Inclure l'historique détaillé
    
    Returns:
        Chemin vers le rapport généré
    """
    try:
        if not equipment_data:
            return "❌ Aucune donnée d'équipement fournie"
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"rapport_equipements_{timestamp}.pdf"
        filepath = f"reports/{filename}"
        
        os.makedirs("reports", exist_ok=True)
        
        if not REPORTLAB_AVAILABLE:
            return "❌ ReportLab non installé. Installez-le avec: pip install reportlab"
        
        doc = SimpleDocTemplate(filepath, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Titre
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER
        )
        story.append(Paragraph("Rapport de Maintenance par Équipement", title_style))
        story.append(Spacer(1, 20))
        
        # Résumé global
        total_equipment = len(equipment_data)
        critical_equipment = len([e for e in equipment_data if e.get('criticality', '').lower() in ['critical', 'critique', 'high']])
        maintenance_due = len([e for e in equipment_data if e.get('maintenance_due', False)])
        
        summary_data = [
            ['Métrique', 'Valeur'],
            ['Total équipements', str(total_equipment)],
            ['Équipements critiques', str(critical_equipment)],
            ['Maintenance due', str(maintenance_due)],
            ['Période analysée', f"{report_period} jours"]
        ]
        
        summary_table = Table(summary_data, colWidths=[2*inch, 1.5*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Détail par équipement
        for equipment in equipment_data:
            story.append(Paragraph(f"Équipement: {equipment.get('id', 'N/A')}", styles['Heading2']))
            story.append(Spacer(1, 12))
            
            # Informations de base
            basic_info = [
                ['Propriété', 'Valeur'],
                ['Description', equipment.get('description', 'N/A')],
                ['Localisation', equipment.get('location', 'N/A')],
                ['Criticité', equipment.get('criticality', 'N/A')],
                ['Statut', equipment.get('status', 'N/A')],
                ['Dernière maintenance', equipment.get('last_maintenance', 'N/A')]
            ]
            
            info_table = Table(basic_info, colWidths=[1.5*inch, 3*inch])
            info_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            
            story.append(info_table)
            story.append(Spacer(1, 15))
            
            # Historique de maintenance (si demandé)
            if include_history and equipment.get('maintenance_history'):
                story.append(Paragraph("Historique de Maintenance", styles['Heading3']))
                story.append(Spacer(1, 8))
                
                history_data = [['Date', 'Type', 'Description', 'Technicien']]
                for maintenance in equipment.get('maintenance_history', [])[:5]:  # Limite à 5 entrées
                    history_data.append([
                        maintenance.get('date', 'N/A'),
                        maintenance.get('type', 'N/A'),
                        maintenance.get('description', 'N/A')[:30] + '...' if len(maintenance.get('description', '')) > 30 else maintenance.get('description', 'N/A'),
                        maintenance.get('technician', 'N/A')
                    ])
                
                history_table = Table(history_data, colWidths=[1*inch, 1*inch, 2.5*inch, 1*inch])
                history_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ]))
                
                story.append(history_table)
            
            story.append(Spacer(1, 20))
        
        # Pied de page
        footer_text = f"Rapport généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        story.append(Paragraph(footer_text, styles['Normal']))
        
        doc.build(story)
        
        return f"✅ Rapport équipements généré avec succès: {filepath}"
        
    except Exception as e:
        return f"❌ Erreur lors de la génération du rapport équipements: {str(e)}"

@tool
async def generate_kpi_dashboard(
    kpi_data: Dict[str, Any],
    dashboard_title: str = "Tableau de Bord KPI Maintenance"
) -> str:
    """
    Génère un tableau de bord KPI pour la maintenance.
    Utilisez cet outil pour créer des rapports de performance.
    
    Args:
        kpi_data: Données des indicateurs de performance
        dashboard_title: Titre du tableau de bord
    
    Returns:
        Chemin vers le tableau de bord généré
    """
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"dashboard_kpi_{timestamp}.pdf"
        filepath = f"reports/{filename}"
        
        os.makedirs("reports", exist_ok=True)
        
        if not REPORTLAB_AVAILABLE:
            return "❌ ReportLab non installé. Installez-le avec: pip install reportlab"
        
        doc = SimpleDocTemplate(filepath, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Titre
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            spaceAfter=30,
            alignment=TA_CENTER
        )
        story.append(Paragraph(dashboard_title, title_style))
        story.append(Spacer(1, 20))
        
        # KPIs principaux
        story.append(Paragraph("Indicateurs de Performance Clés", styles['Heading2']))
        story.append(Spacer(1, 15))
        
        # Tableau des KPIs
        kpi_headers = ['KPI', 'Valeur', 'Objectif', 'Statut']
        kpi_table_data = [kpi_headers]
        
        for kpi_name, kpi_info in kpi_data.items():
            current_value = kpi_info.get('current', 'N/A')
            target_value = kpi_info.get('target', 'N/A')
            status = kpi_info.get('status', 'N/A')
            
            # Couleur selon le statut
            status_color = colors.green if status == 'OK' else colors.red if status == 'KO' else colors.orange
            
            kpi_table_data.append([kpi_name, str(current_value), str(target_value), status])
        
        kpi_table = Table(kpi_table_data, colWidths=[2*inch, 1*inch, 1*inch, 1*inch])
        kpi_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
        ]))
        
        story.append(kpi_table)
        story.append(Spacer(1, 25))
        
        # Résumé et recommandations
        story.append(Paragraph("Résumé et Recommandations", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        # Calcul des statistiques
        total_kpis = len(kpi_data)
        ok_kpis = len([k for k in kpi_data.values() if k.get('status') == 'OK'])
        improvement_needed = total_kpis - ok_kpis
        
        summary_text = f"""
        <b>Résumé:</b><br/>
        • Total KPIs analysés: {total_kpis}<br/>
        • KPIs dans l'objectif: {ok_kpis}<br/>
        • KPIs nécessitant une amélioration: {improvement_needed}<br/>
        • Taux de réussite: {(ok_kpis/total_kpis*100):.1f}%<br/><br/>
        
        <b>Recommandations:</b><br/>
        • Maintenir les bonnes pratiques pour les KPIs verts<br/>
        • Analyser les causes des KPIs rouges et oranges<br/>
        • Mettre en place des actions correctives prioritaires<br/>
        • Planifier des revues régulières des objectifs
        """
        
        story.append(Paragraph(summary_text, styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Pied de page
        footer_text = f"Tableau de bord généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        story.append(Paragraph(footer_text, styles['Normal']))
        
        doc.build(story)
        
        return f"✅ Tableau de bord KPI généré avec succès: {filepath}"
        
    except Exception as e:
        return f"❌ Erreur lors de la génération du tableau de bord: {str(e)}"

@tool
async def export_data_to_excel(
    data: List[Dict[str, Any]],
    sheet_name: str = "Données",
    filename: str = None
) -> str:
    """
    Exporte des données vers un fichier Excel.
    Utilisez cet outil pour créer des fichiers Excel avec les données de maintenance.
    
    Args:
        data: Données à exporter
        sheet_name: Nom de la feuille Excel
        filename: Nom du fichier (optionnel)
    
    Returns:
        Chemin vers le fichier Excel généré
    """
    try:
        # Vérification de la disponibilité d'openpyxl
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            OPENPYXL_AVAILABLE = True
        except ImportError:
            return "❌ OpenPyXL non installé. Installez-le avec: pip install openpyxl"
        
        if not data:
            return "❌ Aucune donnée à exporter"
        
        # Génération du nom de fichier
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"export_donnees_{timestamp}.xlsx"
        
        filepath = f"reports/{filename}"
        os.makedirs("reports", exist_ok=True)
        
        # Création du workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # En-têtes
        if data:
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
        
        # Données
        for row, item in enumerate(data, 2):
            for col, header in enumerate(headers, 1):
                value = item.get(header, '')
                ws.cell(row=row, column=col, value=value)
        
        # Ajustement automatique des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Sauvegarde
        wb.save(filepath)
        
        return f"✅ Fichier Excel généré avec succès: {filepath}"
        
    except Exception as e:
        return f"❌ Erreur lors de l'export Excel: {str(e)}" 